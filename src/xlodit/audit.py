from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string

from xlodit.backend import BackendError, recalculate_with_libreoffice, resolve_backend
from xlodit.models import AuditIssue, AuditReport, BackendInfo
from xlodit.rules import (
    formula_has_error_token,
    is_formula_cell,
    is_formula_error_token,
    numeric_text_issues,
    rule_wb001,
    structural_issues,
)
from xlodit.traceback_engine import build_traceback


class AuditRuntimeError(RuntimeError):
    pass


def _cell_sort_key(cell: str | None) -> tuple[int, int]:
    if not cell:
        return (10**9, 10**9)
    try:
        col, row = coordinate_from_string(cell)
        return (int(row), int(column_index_from_string(col)))
    except Exception:
        return (10**9, 10**9)


def _issue_sort_key(issue: AuditIssue) -> tuple[int, str, tuple[int, int], str]:
    severity_rank = 0 if issue.severity == "error" else 1
    return (severity_rank, issue.sheet or "", _cell_sort_key(issue.cell), issue.rule_id)


def _formula_issues(
    source_wb,
    values_wb,
    workbook_path: Path,
    max_traceback_depth: int,
) -> list[AuditIssue]:
    issues: list[AuditIssue] = []
    for sheet in source_wb.worksheets:
        values_sheet = values_wb[sheet.title] if sheet.title in values_wb else None
        for row in sheet.iter_rows():
            for cell in row:
                if not is_formula_cell(cell.value):
                    continue
                formula = str(cell.value)
                value_cell = (
                    values_sheet[cell.coordinate] if values_sheet is not None else None
                )
                value = value_cell.value if value_cell is not None else None
                if is_formula_error_token(value) or formula_has_error_token(formula):
                    issue = AuditIssue(
                        severity="error",
                        kind="formula_error",
                        rule_id="F001",
                        message="Formula evaluation error detected",
                        workbook=str(workbook_path),
                        sheet=sheet.title,
                        cell=cell.coordinate,
                        formula=formula,
                        value=value
                        if isinstance(value, (str, int, float))
                        else str(value),
                    )
                    issue.traceback = build_traceback(
                        source_wb=source_wb,
                        values_wb=values_wb,
                        sheet=sheet.title,
                        cell=cell.coordinate,
                        max_depth=max_traceback_depth,
                    )
                    issues.append(issue)
                    continue
                if value_cell is not None and value_cell.data_type == "e":
                    issues.append(
                        AuditIssue(
                            severity="error",
                            kind="formula_error_datatype",
                            rule_id="F002",
                            message="Formula result has error datatype",
                            workbook=str(workbook_path),
                            sheet=sheet.title,
                            cell=cell.coordinate,
                            formula=formula,
                            value=str(value),
                        )
                    )
    return issues


def _report(
    backend: BackendInfo,
    issues: list[AuditIssue],
    audit_completed: bool,
) -> AuditReport:
    issues = sorted(issues, key=_issue_sort_key)
    errors = sum(1 for i in issues if i.severity == "error")
    warnings = sum(1 for i in issues if i.severity == "warning")
    summary = {"errors": errors, "warnings": warnings, "total": errors + warnings}
    return AuditReport(
        ok=(errors == 0 and audit_completed),
        audit_completed=audit_completed,
        backend=backend,
        summary=summary,
        issues=issues,
    )


def audit_workbook(
    path: str, backend: str = "auto", max_traceback_depth: int = 10
) -> AuditReport:
    workbook_path = Path(path)
    backend_info = BackendInfo(requested=backend, used="none", recalculated=False)

    if workbook_path.suffix.lower() != ".xlsx":
        issue = AuditIssue(
            severity="error",
            kind="workbook_unreadable",
            rule_id="WB001",
            message="Unsupported file type; only .xlsx is supported",
            workbook=str(workbook_path),
        )
        return _report(backend_info, [issue], audit_completed=False)

    try:
        source_wb = load_workbook(workbook_path, data_only=False)
    except Exception as exc:
        return _report(
            backend_info,
            [rule_wb001(workbook_path, f"Unable to read workbook: {exc}")],
            False,
        )

    try:
        used_backend = resolve_backend(backend)
    except BackendError as exc:
        issue = AuditIssue(
            severity="error",
            kind="backend_unavailable",
            rule_id="BE001",
            message=str(exc),
            workbook=str(workbook_path),
            backend="libreoffice",
        )
        return _report(backend_info, [issue], audit_completed=False)

    backend_info.used = used_backend
    values_path = workbook_path
    if used_backend == "libreoffice":
        try:
            values_path = recalculate_with_libreoffice(workbook_path)
            backend_info.recalculated = True
        except BackendError as exc:
            issue = AuditIssue(
                severity="error",
                kind="backend_runtime_failure",
                rule_id="BE002",
                message=str(exc),
                workbook=str(workbook_path),
                backend="libreoffice",
            )
            return _report(backend_info, [issue], audit_completed=False)

    try:
        values_wb = load_workbook(values_path, data_only=True)
    except Exception as exc:
        return _report(
            backend_info,
            [rule_wb001(workbook_path, f"Unable to read evaluated workbook: {exc}")],
            False,
        )

    issues: list[AuditIssue] = []
    issues.extend(structural_issues(source_wb, workbook_path))
    issues.extend(
        _formula_issues(source_wb, values_wb, workbook_path, max_traceback_depth)
    )
    issues.extend(numeric_text_issues(values_wb, workbook_path))
    return _report(backend_info, issues, audit_completed=True)
