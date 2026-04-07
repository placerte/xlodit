from openpyxl import load_workbook

from xlodit.traceback_engine import build_traceback


def test_traceback_cycle_is_partial(fixture_dir) -> None:
    path = fixture_dir / "cycle.xlsx"
    source = load_workbook(path, data_only=False)
    values = load_workbook(path, data_only=True)
    result = build_traceback(source, values, sheet="Sheet1", cell="A1", max_depth=10)
    assert result.status == "partial"
    assert result.cycle_detected is True


def test_traceback_depth_limit_marks_partial(fixture_dir) -> None:
    path = fixture_dir / "formula_chain.xlsx"
    source = load_workbook(path, data_only=False)
    values = load_workbook(path, data_only=True)
    result = build_traceback(source, values, sheet="Sheet1", cell="C1", max_depth=1)
    assert result.status == "partial"
    assert result.max_depth_reached is True
