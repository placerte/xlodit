from __future__ import annotations

import sys
from pathlib import Path

import pytest
from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "src"
if str(SRC) not in sys.path:
    sys.path.insert(0, str(SRC))


def _save(workbook: Workbook, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def _fixture_valid(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = 1
    ws["B1"] = 2
    _save(wb, path)


def _fixture_formula_error_simple(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "=#DIV/0!"
    _save(wb, path)


def _fixture_formula_chain(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "=#DIV/0!"
    ws["B1"] = "=A1"
    ws["C1"] = "=B1"
    _save(wb, path)


def _fixture_numeric_text(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "12,34"
    ws["A2"] = "56.78"
    _save(wb, path)


def _fixture_missing_reference(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "=Z99"
    _save(wb, path)


def _fixture_cycle(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "=B1"
    ws["B1"] = "=A1"
    _save(wb, path)


def _fixture_mixed(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Main"
    ws["A1"] = "=#REF!"
    ws["B1"] = "7,01"
    ws["C1"] = "=A1"
    _save(wb, path)


@pytest.fixture(scope="session")
def fixture_dir(tmp_path_factory: pytest.TempPathFactory) -> Path:
    root = tmp_path_factory.mktemp("xlodit-fixtures")
    files = {
        "valid.xlsx": _fixture_valid,
        "formula_error_simple.xlsx": _fixture_formula_error_simple,
        "formula_chain.xlsx": _fixture_formula_chain,
        "numeric_text.xlsx": _fixture_numeric_text,
        "missing_reference.xlsx": _fixture_missing_reference,
        "cycle.xlsx": _fixture_cycle,
        "mixed.xlsx": _fixture_mixed,
    }
    for name, builder in files.items():
        builder(root / name)
    return root
